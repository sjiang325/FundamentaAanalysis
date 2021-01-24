import os
import pandas as pd
import pandas_datareader as web
from pandas.tseries.offsets import MonthEnd
import numpy as np
import openpyxl

version = 2  # last modified 2020-11-30


def my_own_parent_directory():
    p = os.path.dirname(__file__)
    return p


def excel_add_sheet(df, workbook_name, sheet_name, folder_path=None, **kwargs):
    if folder_path is None:
        folder_path = my_own_parent_directory()

    if workbook_name[-5:].lower() != '.xlsx':
        workbook_name = workbook_name + '.xlsx'

    full_path = os.path.join(folder_path, workbook_name)

    if not os.path.exists(full_path):
        with pd.ExcelWriter(full_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name=sheet_name, **kwargs)
    else:
        # print('file exists {}'.format(full_path))
        workbook = openpyxl.load_workbook(full_path)
        if sheet_name not in workbook.sheetnames:
            with pd.ExcelWriter(full_path, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=sheet_name, **kwargs)
        else:
            if len(workbook.sheetnames) == 0:
                print('*** Some kind of error happened ***')
            elif len(workbook.sheetnames) == 1:
                # print('only one sheet')
                workbook.close()
                os.remove(full_path)
                with pd.ExcelWriter(full_path, engine='openpyxl', mode='w') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, **kwargs)
            else:
                sheet = workbook.get_sheet_by_name(sheet_name)
                # print('got sheet')
                workbook.remove_sheet(sheet)
                workbook.save(full_path)
                workbook.close()
                with pd.ExcelWriter(full_path, engine='openpyxl', mode='a') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, **kwargs)


def get_yahoo_returns_monthly_dbstyle(asset_id, date_beg, date_end):

    df_ret = get_yahoo_returns_monthly(asset_id, date_beg, date_end)

    df_ret['Ticker'] = asset_id
    df_ret['Date'] = df_ret.index

    df_ret = df_ret[['Ticker','Date','MthlyRet']]
    df_ret = df_ret.reset_index(drop=True)

    return df_ret



def get_yahoo_returns_monthly(asset_id, date_beg, date_end):
    """
      Creates a dataframe with monthly returns from yahoo

      Args:
          asset_id (str): valid id in finance.yahoo.com
          date_beg (str): monthly date in form 'yyyy-mm' indicating first month
          date_end (str): monthly date in form 'yyyy-mm' indicating last month

      Returns:
          Pandas DataFrame: pandas dataframe with monthly returns

    """

    # converts date to last day of month represented by date
    date_beg = pd.to_datetime(date_beg) + MonthEnd(0)
    date_end = pd.to_datetime(date_end) + MonthEnd(0)

    # IMPORTANT: get price data for prior month, so return can be calculated for
    # first month, which of course, needs price for the month-end prior.
    # create new variable used in this function for the start date
    start_date = date_beg + pd.DateOffset(months=-1)  # go back one month

    # get prices
    df_prc = get_yahoo_price_data_monthly(asset_id, start_date, date_end)

    # create an empty data frame with same dates as port
    df_ret = pd.DataFrame(index=df_prc.index)

    # calculate percent return for each and add column to the return dataframe
    # with column name equal to the id of the security
    df_ret = df_prc[['Adj Close']].pct_change() * 100
    #
    # we have an extra row that came from the prices.
    # so drop the first row which is NaN
    df_ret = df_ret[1:]

    # rename column to ret
    df_ret = df_ret.rename(columns={"Adj Close": "MthlyRet"})

    # convert index to pandas times series period monthly
    mthly_dates = df_ret.index.strftime('%Y-%m')
    periods = list(map(pd.Period, mthly_dates))

    df_ret['Date'] = periods
    df_ret = df_ret.set_index('Date')

    return df_ret
    
    
def get_yahoo_price_data_monthly(asset_id, date_beg, date_end):
    """
    Creates a dataframe with month-end prices from yahoo.
    This is done by getting daily prices for the dates spanned and then
    filtering to get just the month-end dates.

    IMPORTANT: this function calls a function that gets daily prices. That function
    has an option called "fill_forward". If the option is set to "True," any of
    the days in the daily dates over the date range that is a non-trading day
    pulls or "fills forward" prices from the most recent price available.
    This is to address month-ends that we will want that may be on weekends, holidays, etc.
    The default setting in the daily price retrieval function that is used
    for fill_forward function is "False," so when it is used here, we want
    to set it to True, to fill_forward

    Args:
        asset_id (str): valid id in finance.yahoo.com
        date_beg (str): monthly date in form 'yyyy-mm' indicating first month
        date_end (str): monthly date in form 'yyyy-mm' indicating last month

    Returns:
        Pandas DataFrame: pandas dataframe with month-end prices

    """

    # converts date to last day of month represented by date
    date_beg = pd.to_datetime(date_beg) + MonthEnd(0)
    date_end = pd.to_datetime(date_end) + MonthEnd(0)

    df_price_daily = get_yahoo_price_data_daily(asset_id, date_beg, date_end, fill_forward=True)

    # Create month end dates
    month_end_dates = pd.date_range(date_beg, date_end, freq='M')

    # Create empty data frame with index the daily dates
    df_price = pd.DataFrame(index=month_end_dates)
    df_price = df_price.merge(df_price_daily['Adj Close'], left_index=True, right_index=True)

    return df_price


def get_yahoo_price_data_daily(asset_id, date_beg, date_end, fill_forward=False):
    # establish the daily begin and end dates
    beg_date = pd.to_datetime(date_beg)
    end_date = pd.to_datetime(date_end)

    # NOTE: going to add one additional month of daily data for data retrieval.
    # This is in case beginning date is a non-trading day, in which case
    start_date = date_beg + pd.DateOffset(months=-1)

    # Pulls historical price data for given ticker
    # price_data_yahoo = yf.Ticker(asset_id).history(start=start_date, end=end_date, interval="1d")
    price_data_yahoo = web.DataReader(asset_id, 'yahoo', start_date, end_date)

    # Creates vector containing all daily dates in period
    daily_dates = pd.date_range(start_date, end_date)

    # Create empty data frame with index the daily dates
    df_price = pd.DataFrame(index=daily_dates)

    # Merges price data with full set of dates to account for missing dates.
    df_price = df_price.merge(right=price_data_yahoo['Adj Close'], how='left', left_index=True, right_index=True)

    # if "Fill forward" any missing data from prior data
    if fill_forward:
        df_price = df_price.fillna(method='ffill')

    # now only include data in requested range
    df_price = df_price[df_price.index >= date_beg]

    return df_price


if __name__ == '__main__':

    df_sample = pd.DataFrame(data=[['ABC',3.2],['XYZ',1.7]], columns=['ticker','price'])
    print(df_sample)
    excel_add_sheet(df_sample, 'sample.xlsx', 'holdings3', index=None)